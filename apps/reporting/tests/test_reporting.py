"""Tests des services de génération de rapports (Phase 13).

Vérifie que les trois formats (PDF, Excel, CSV) se génèrent correctement
sans erreur, retournent un contenu non vide et le bon type de fichier.
"""
import io

import pytest

from apps.reporting.services.csv_export import exporter_csv, HEADERS
from apps.reporting.services.excel_export import exporter_excel
from apps.reporting.services.pdf_export import exporter_pdf


# Marqueur `slow` : ces tests nécessitent une base de données.
pytestmark = pytest.mark.django_db


class TestExportCSV:
    """Tests de l'export CSV."""

    def test_csv_headers_complets(self):
        """L'en-tête CSV doit contenir les 18 colonnes attendues."""
        assert len(HEADERS) == 18
        assert "matricule" in HEADERS
        assert "score_risque" in HEADERS
        assert "classification_risque" in HEADERS

    def test_csv_generation_vide(self):
        """L'export CSV doit fonctionner même sans données (juste l'en-tête)."""
        buffer = exporter_csv()
        content = buffer.getvalue()
        lines = content.strip().split("\n")
        # Au moins la ligne d'en-tête
        assert len(lines) >= 1
        assert "matricule" in lines[0]

    def test_csv_type_retour(self):
        """exporter_csv doit retourner un StringIO."""
        buffer = exporter_csv()
        assert isinstance(buffer, io.StringIO)

    def test_csv_avec_donnees(self, db, etudiant_avec_indicateur):
        """L'export CSV doit contenir une ligne de données par étudiant."""
        buffer = exporter_csv()
        content = buffer.getvalue()
        lines = content.strip().split("\n")
        # En-tête + 1 ligne de données
        assert len(lines) >= 2
        assert etudiant_avec_indicateur.matricule in content

    def test_csv_filtre_filiere(self, db, etudiant_avec_indicateur):
        """Le filtre par filière doit fonctionner."""
        filiere = etudiant_avec_indicateur.filiere
        buffer = exporter_csv(filiere=filiere)
        content = buffer.getvalue()
        assert etudiant_avec_indicateur.matricule in content


class TestExportExcel:
    """Tests de l'export Excel."""

    def test_excel_generation_vide(self):
        """L'export Excel doit fonctionner même sans données."""
        buffer = exporter_excel()
        data = buffer.getvalue()
        # Signature d'un fichier XLSX (format ZIP)
        assert data[:4] == b"PK\x03\x04"
        assert len(data) > 0

    def test_excel_type_retour(self):
        """exporter_excel doit retourner un BytesIO."""
        buffer = exporter_excel()
        assert isinstance(buffer, io.BytesIO)

    def test_excel_avec_donnees(self, db, etudiant_avec_indicateur):
        """L'export Excel doit générer un fichier valide avec données."""
        buffer = exporter_excel()
        data = buffer.getvalue()
        assert data[:4] == b"PK\x03\x04"
        assert len(data) > 1000  # Au moins 1 Ko

    def test_excel_3_feuilles(self, db, etudiant_avec_indicateur):
        """Le classeur doit contenir 3 feuilles : Synthèse, Résultats, Alertes."""
        from openpyxl import load_workbook
        buffer = exporter_excel()
        buffer.seek(0)
        wb = load_workbook(buffer)
        assert "Synthèse" in wb.sheetnames
        assert "Résultats" in wb.sheetnames
        assert "Alertes" in wb.sheetnames


class TestExportPDF:
    """Tests de l'export PDF."""

    def test_pdf_generation_vide(self):
        """L'export PDF doit fonctionner même sans données."""
        buffer = exporter_pdf()
        data = buffer.getvalue()
        # Signature d'un fichier PDF
        assert data[:5] == b"%PDF-"
        assert len(data) > 0

    def test_pdf_type_retour(self):
        """exporter_pdf doit retourner un BytesIO."""
        buffer = exporter_pdf()
        assert isinstance(buffer, io.BytesIO)

    def test_pdf_avec_donnees(self, db, etudiant_avec_indicateur):
        """L'export PDF doit générer un fichier valide avec données."""
        buffer = exporter_pdf()
        data = buffer.getvalue()
        assert data[:5] == b"%PDF-"
        # Fin du fichier PDF
        assert b"%%EOF" in data[-1024:]
        assert len(data) > 2000

    def test_pdf_filtre_niveau(self, db, etudiant_avec_indicateur):
        """Le filtre par niveau doit fonctionner."""
        niveau = etudiant_avec_indicateur.niveau
        buffer = exporter_pdf(niveau=niveau)
        data = buffer.getvalue()
        assert data[:5] == b"%PDF-"
