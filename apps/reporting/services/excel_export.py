"""Export Excel des indicateurs académiques (BF08).

Génère un classeur Excel avec :
  - Feuille « Synthèse » : indicateurs par étudiant
  - Feuille « Résultats » : résultats détaillés par UE et semestre
  - Feuille « Alertes » : alertes actives

Utilise openpyxl (sans dépendance Excel installé).
"""
from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.accounts.models import ProfilEtudiant
from apps.academics.models import ResultatAcademique
from apps.analytics.models import IndicateurAcademique
from apps.alerts.models import Alerte

# Styles réutilisables
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_RISQUE_FILLS = {
    "Faible": PatternFill(start_color="38A169", end_color="38A169", fill_type="solid"),
    "Modere": PatternFill(start_color="D69E2E", end_color="D69E2E", fill_type="solid"),
    "Eleve": PatternFill(start_color="C53030", end_color="C53030", fill_type="solid"),
}


def _style_header(ws, row, max_col):
    """Applique le style d'en-tête à une ligne."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER


def _auto_width(ws, max_col, max_row, min_width=10, max_width=40):
    """Ajuste la largeur des colonnes automatiquement."""
    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, min_width), max_width)


def exporter_excel(
    filiere: Optional[str] = None,
    niveau: Optional[str] = None,
) -> io.BytesIO:
    """Génère un export Excel complet (synthèse + résultats + alertes).

    Args:
        filiere: filtre par filière (None = toutes).
        niveau: filtre par niveau LMD (None = tous).

    Returns:
        BytesIO contenant le classeur Excel.
    """
    wb = Workbook()

    # --- Feuille Synthèse ---
    ws_synthese = wb.active
    ws_synthese.title = "Synthèse"
    _remplir_synthese(ws_synthese, filiere, niveau)

    # --- Feuille Résultats ---
    ws_resultats = wb.create_sheet("Résultats")
    _remplir_resultats(ws_resultats, filiere, niveau)

    # --- Feuille Alertes ---
    ws_alertes = wb.create_sheet("Alertes")
    _remplir_alertes(ws_alertes, filiere, niveau)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _remplir_synthese(ws, filiere=None, niveau=None):
    """Remplit la feuille de synthèse des indicateurs."""
    headers = [
        "Matricule", "Nom", "Prénom", "Niveau", "Filière", "Promotion",
        "Année scolaire", "Moy. générale", "Score risque", "Classification",
        "Progression %", "Crédits acquis", "Crédits total",
        "UE échec", "UE total", "Absentéisme %", "Sem. restants", "Proj. diplomation",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    qs = _queryset_indicateurs(filiere, niveau)
    row = 2
    for indic in qs:
        etu = indic.etudiant
        user = etu.utilisateur
        data = [
            etu.matricule,
            user.last_name,
            user.first_name,
            etu.niveau,
            etu.filiere,
            etu.promotion,
            etu.annee_scolaire,
            _f(indic.moyenne_generale),
            _f(indic.score_risque),
            indic.classification_risque or "",
            _f(indic.taux_progression),
            indic.credits_acquis,
            indic.credits_total,
            indic.ues_echec,
            indic.ues_total,
            _f(indic.taux_absenteisme),
            _f(indic.semestres_restants),
            indic.proj_diplomation.isoformat() if indic.proj_diplomation else "",
        ]
        ws.append(data)
        # Bordures
        for col in range(1, len(data) + 1):
            ws.cell(row=row, column=col).border = _BORDER
        # Couleur classification risque (colonne 10)
        cls = indic.classification_risque
        if cls in _RISQUE_FILLS:
            cell = ws.cell(row=row, column=10)
            cell.fill = _RISQUE_FILLS[cls]
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        row += 1

    _auto_width(ws, len(headers), row - 1)


def _remplir_resultats(ws, filiere=None, niveau=None):
    """Remplit la feuille des résultats détaillés par UE."""
    headers = [
        "Matricule", "Nom", "Code UE", "Intitulé UE", "Note", "Crédits",
        "Validé", "Session", "Semestre", "Année scolaire",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    qs = ResultatAcademique.objects.select_related(
        "etudiant", "etudiant__utilisateur", "ue", "semestre"
    ).all()
    if filiere:
        qs = qs.filter(etudiant__filiere=filiere)
    if niveau:
        qs = qs.filter(etudiant__niveau=niveau)

    row = 2
    for res in qs[:5000]:  # Limite pour éviter un fichier trop lourd
        etu = res.etudiant
        user = etu.utilisateur
        data = [
            etu.matricule,
            user.get_full_name() or user.username,
            res.ue.code,
            res.ue.intitule,
            _f(res.note),
            res.credits,
            "Oui" if res.valide else "Non",
            res.session,
            str(res.semestre),
            res.semestre.annee_scolaire if hasattr(res.semestre, "annee_scolaire") else "",
        ]
        ws.append(data)
        for col in range(1, len(data) + 1):
            ws.cell(row=row, column=col).border = _BORDER
        row += 1

    _auto_width(ws, len(headers), row - 1)


def _remplir_alertes(ws, filiere=None, niveau=None):
    """Remplit la feuille des alertes actives."""
    headers = [
        "Matricule", "Nom", "Type", "Niveau risque", "Score",
        "Message", "Statut", "Date création", "Priorité reco", "Recommandation",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    qs = Alerte.objects.select_related(
        "etudiant", "etudiant__utilisateur"
    ).prefetch_related("recommandations").order_by("-created_at")
    if filiere:
        qs = qs.filter(etudiant__filiere=filiere)
    if niveau:
        qs = qs.filter(etudiant__niveau=niveau)

    row = 2
    for alerte in qs[:2000]:
        etu = alerte.etudiant
        # Recommandation principale
        reco_desc = ""
        reco_prio = ""
        if alerte.recommandations.exists():
            reco = alerte.recommandations.first()
            reco_desc = reco.description[:200]
            reco_prio = reco.priorite
        data = [
            etu.matricule,
            etu.utilisateur.get_full_name() or etu.utilisateur.username,
            alerte.get_type_display() if hasattr(alerte, "get_type_display") else alerte.type,
            alerte.get_niveau_risque_display() if hasattr(alerte, "get_niveau_risque_display") else alerte.niveau_risque,
            _f(alerte.score_risque),
            alerte.message[:300] if alerte.message else "",
            alerte.get_statut_display() if hasattr(alerte, "get_statut_display") else alerte.statut,
            alerte.created_at.strftime("%d/%m/%Y %H:%M") if alerte.created_at else "",
            reco_prio,
            reco_desc,
        ]
        ws.append(data)
        for col in range(1, len(data) + 1):
            ws.cell(row=row, column=col).border = _BORDER
        # Couleur niveau de risque (colonne 4)
        niveau_r = alerte.niveau_risque
        if niveau_r in _RISQUE_FILLS:
            ws.cell(row=row, column=4).fill = _RISQUE_FILLS[niveau_r]
        row += 1

    _auto_width(ws, len(headers), row - 1)


def _queryset_indicateurs(filiere=None, niveau=None):
    """QuerySet filtré des indicateurs globaux (semestre=None)."""
    qs = IndicateurAcademique.objects.filter(
        semestre=None
    ).select_related("etudiant", "etudiant__utilisateur")
    if filiere:
        qs = qs.filter(etudiant__filiere=filiere)
    if niveau:
        qs = qs.filter(etudiant__niveau=niveau)
    return qs


def _f(valeur) -> str:
    """Formate une valeur numérique pour Excel."""
    if valeur is None:
        return ""
    if isinstance(valeur, float):
        return round(valeur, 2)
    return valeur
